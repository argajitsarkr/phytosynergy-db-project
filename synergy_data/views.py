import csv
import io
import json
import logging
import os
import re
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from .forms import (
    BULK_CSV_COLUMNS,
    COLUMN_MAP,
    BulkCSVUploadForm,
    SynergyEntryForm,
    _canonical_header,
)
from .models import (
    Antibiotic,
    AntibioticClass,
    Pathogen,
    Phytochemical,
    Plant,
    Source,
    SynergyExperiment,
)
from .pubchem_utils import enrich_phytochemical
from . import similarity

logger = logging.getLogger(__name__)


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def parse_pathogen_name(full_name):
    """
    Parse a pathogen full name into (genus, species, strain).

    Examples:
        "Pseudomonas aeruginosa MTCC 2488" -> ("Pseudomonas", "aeruginosa", "MTCC 2488")
        "Staphylococcus aureus"             -> ("Staphylococcus", "aureus", None)
        "MRSA"                              -> ("MRSA", "", None)
    """
    parts = full_name.strip().split()
    if len(parts) >= 2:
        genus = parts[0]
        species = parts[1]
        strain = ' '.join(parts[2:]) if len(parts) > 2 else None
    elif len(parts) == 1:
        genus = parts[0]
        species = ''
        strain = None
    else:
        raise ValueError("Pathogen name cannot be empty")
    return genus, species, strain or None


def auto_calculate_fic(mic_phyto_alone, mic_abx_alone, mic_phyto_in_combo, mic_abx_in_combo):
    """
    FIC = (MIC_phyto_combo / MIC_phyto_alone) + (MIC_abx_combo / MIC_abx_alone)
    Returns None if any required value is missing or zero.
    """
    values = [mic_phyto_alone, mic_abx_alone, mic_phyto_in_combo, mic_abx_in_combo]
    if all(v is not None and v > 0 for v in values):
        return (mic_phyto_in_combo / mic_phyto_alone) + (mic_abx_in_combo / mic_abx_alone)
    return None


def auto_interpret_fic(fic_index):
    """Derive interpretation from FIC index value."""
    if fic_index is None:
        return None
    if fic_index <= Decimal('0.5'):
        return 'Synergy'
    elif fic_index <= Decimal('1.0'):
        return 'Additive'
    elif fic_index <= Decimal('4.0'):
        return 'Indifference'
    else:
        return 'Antagonism'


def get_or_create_case_insensitive(model, field_name, value):
    """
    Case-insensitive get_or_create for models with unique text fields.
    Handles race conditions with IntegrityError fallback.
    """
    lookup = {f'{field_name}__iexact': value}
    try:
        return model.objects.get(**lookup)
    except model.DoesNotExist:
        try:
            return model.objects.create(**{field_name: value})
        except IntegrityError:
            return model.objects.get(**lookup)


# Gram stain by genus for the common antibacterial test organisms. Used to
# auto-fill Pathogen.gram_stain when a curator doesn't supply it explicitly.
GRAM_STAIN_BY_GENUS = {
    # Gram-positive
    'staphylococcus': 'Gram-positive',
    'streptococcus': 'Gram-positive',
    'enterococcus': 'Gram-positive',
    'bacillus': 'Gram-positive',
    'listeria': 'Gram-positive',
    'clostridium': 'Gram-positive',
    'clostridioides': 'Gram-positive',
    'corynebacterium': 'Gram-positive',
    'mycobacterium': 'Gram-positive',
    'micrococcus': 'Gram-positive',
    'lactobacillus': 'Gram-positive',
    # Gram-negative
    'escherichia': 'Gram-negative',
    'klebsiella': 'Gram-negative',
    'pseudomonas': 'Gram-negative',
    'acinetobacter': 'Gram-negative',
    'enterobacter': 'Gram-negative',
    'salmonella': 'Gram-negative',
    'shigella': 'Gram-negative',
    'proteus': 'Gram-negative',
    'serratia': 'Gram-negative',
    'neisseria': 'Gram-negative',
    'haemophilus': 'Gram-negative',
    'helicobacter': 'Gram-negative',
    'vibrio': 'Gram-negative',
    'campylobacter': 'Gram-negative',
    'bacteroides': 'Gram-negative',
    'citrobacter': 'Gram-negative',
    'morganella': 'Gram-negative',
    'stenotrophomonas': 'Gram-negative',
    'burkholderia': 'Gram-negative',
}


def derive_gram_stain(genus):
    """Best-effort Gram stain from a genus name. Returns None if unknown."""
    if not genus:
        return None
    return GRAM_STAIN_BY_GENUS.get(genus.strip().lower())


def resolve_pathogen(full_name, gram_stain=None):
    """get_or_create a Pathogen from a full name, filling gram_stain if blank.

    An explicit gram_stain wins; otherwise we derive it from the genus.
    """
    genus, species, strain = parse_pathogen_name(full_name)
    pathogen, _ = Pathogen.objects.get_or_create(
        genus=genus, species=species, strain=strain
    )
    if not pathogen.gram_stain:
        resolved = (gram_stain or '').strip() or derive_gram_stain(genus)
        if resolved:
            pathogen.gram_stain = resolved
            pathogen.save(update_fields=['gram_stain'])
    return pathogen


def resolve_antibiotic(name, class_name=None):
    """get_or_create an Antibiotic (case-insensitive) and link its class.

    The AntibioticClass FK is only filled when currently blank, so we never
    clobber a class a curator set deliberately.
    """
    antibiotic = get_or_create_case_insensitive(Antibiotic, 'antibiotic_name', name)
    class_name = (class_name or '').strip()
    if class_name and antibiotic.antibiotic_class is None:
        antibiotic.antibiotic_class = get_or_create_case_insensitive(
            AntibioticClass, 'class_name', class_name
        )
        antibiotic.save(update_fields=['antibiotic_class'])
    return antibiotic


def link_plant_source(plant_name, phytochemical):
    """get_or_create a Plant and associate the phytochemical with it (M2M)."""
    plant_name = (plant_name or '').strip()
    if not plant_name:
        return None
    plant = get_or_create_case_insensitive(Plant, 'scientific_name', plant_name)
    plant.phytochemicals.add(phytochemical)
    return plant


def _normalize_doi(doi):
    """Normalize DOI for case-insensitive comparison: strip URL prefix, lowercase, trim."""
    if not doi:
        return ''
    return re.sub(r'^https?://(dx\.)?doi\.org/', '', str(doi).strip().lower())


def _lookup_existing_paper(doi=None, pmid=None, title=None):
    """Return the first Source whose DOI / PMID / title matches the given paper, or None.

    DOI match is case-insensitive and ignores URL prefixes; title is iexact.
    Used to warn curators when a paper they're submitting is already in the DB.
    """
    ndoi = _normalize_doi(doi)
    if ndoi:
        for s in Source.objects.exclude(doi__isnull=True).exclude(doi=''):
            if _normalize_doi(s.doi) == ndoi:
                return s
    if pmid:
        s = Source.objects.filter(pmid=str(pmid).strip()).first()
        if s:
            return s
    if title:
        s = Source.objects.filter(article_title__iexact=str(title).strip()).first()
        if s:
            return s
    return None


def _apply_search_filters(results, request):
    """Apply all search/filter parameters from the request to the queryset."""
    query = request.GET.get('query')
    pathogen_id = request.GET.get('pathogen')
    antibiotic_id = request.GET.get('antibiotic')
    phytochemical_id = request.GET.get('phytochemical')
    mechanism = request.GET.get('mechanism')
    interpretation = request.GET.get('interpretation')
    eskape = request.GET.get('eskape')
    chemical_class = request.GET.get('chemical_class')

    if query:
        results = results.filter(
            Q(phytochemical__compound_name__icontains=query) |
            Q(antibiotic__antibiotic_name__icontains=query) |
            Q(pathogen__genus__icontains=query) |
            Q(pathogen__species__icontains=query)
        )

    if pathogen_id:
        results = results.filter(pathogen_id=pathogen_id)

    if antibiotic_id:
        results = results.filter(antibiotic_id=antibiotic_id)

    if phytochemical_id:
        results = results.filter(phytochemical_id=phytochemical_id)

    if mechanism:
        results = results.filter(moa_observed__icontains=mechanism)

    if interpretation:
        results = results.filter(interpretation=interpretation)

    if eskape:
        results = results.filter(pathogen__genus__iexact=eskape)

    if chemical_class:
        results = results.filter(phytochemical__chemical_class__iexact=chemical_class)

    return results


# ==============================================================================
# HOME PAGE
# ==============================================================================

def home_page(request):
    synergy_entries_count = SynergyExperiment.objects.count()
    phytochemical_count = Phytochemical.objects.count()
    antibiotic_count = Antibiotic.objects.count()
    source_count = Source.objects.count()
    eskape_pathogen_count = 6
    synergy_confirmed_count = SynergyExperiment.objects.filter(
        interpretation='Synergy'
    ).count()

    # ESKAPE pathogen data with per-genus experiment counts
    eskape_data = [
        {'letter': 'E', 'name': 'Enterococcus faecium', 'genus': 'Enterococcus', 'gram': 'Gram-positive'},
        {'letter': 'S', 'name': 'Staphylococcus aureus', 'genus': 'Staphylococcus', 'gram': 'Gram-positive'},
        {'letter': 'K', 'name': 'Klebsiella pneumoniae', 'genus': 'Klebsiella', 'gram': 'Gram-negative'},
        {'letter': 'A', 'name': 'Acinetobacter baumannii', 'genus': 'Acinetobacter', 'gram': 'Gram-negative'},
        {'letter': 'P', 'name': 'Pseudomonas aeruginosa', 'genus': 'Pseudomonas', 'gram': 'Gram-negative'},
        {'letter': 'E', 'name': 'Enterobacter spp.', 'genus': 'Enterobacter', 'gram': 'Gram-negative'},
    ]
    for p in eskape_data:
        p['count'] = SynergyExperiment.objects.filter(pathogen__genus=p['genus']).count()

    # Recent entries for the homepage
    recent_entries = SynergyExperiment.objects.select_related(
        'phytochemical', 'antibiotic', 'pathogen', 'source'
    ).order_by('-id')[:5]

    # Synergy share for stats band
    if synergy_entries_count > 0:
        synergy_share = round(100.0 * synergy_confirmed_count / synergy_entries_count)
    else:
        synergy_share = 0

    # Distinct journal names for the "Data provenance" marquee on the home page.
    # Pulled live from the curated sources so the strip always reflects the DB.
    # De-duplicated case-insensitively, then split into two rows for the two
    # opposite-scrolling tracks. Only shown when there are enough to scroll.
    seen_journals = set()
    journals = []
    for raw in Source.objects.exclude(journal__isnull=True).exclude(
        journal__exact=''
    ).values_list('journal', flat=True):
        name = (raw or '').strip()
        key = name.lower()
        if name and key not in seen_journals:
            seen_journals.add(key)
            journals.append(name)
    journals.sort(key=str.lower)
    if len(journals) >= 6:
        half = (len(journals) + 1) // 2
        journals_row1 = journals[:half]
        journals_row2 = journals[half:]
    else:
        journals_row1 = []
        journals_row2 = []

    # FAQ entries (also emitted as schema.org JSON-LD in template)
    faq_data = [
        {'q': 'What is the FIC index?',
         'a': 'The Fractional Inhibitory Concentration (FIC) index quantifies how two '
              'antimicrobials interact. Values up to 0.5 indicate synergy, up to 1.0 '
              'additive, up to 4.0 indifference, and above 4.0 antagonism.'},
        {'q': 'How are entries curated?',
         'a': 'Every record is extracted from a peer-reviewed publication with DOI or '
              'PMID. We capture MIC values for each compound alone and in combination, '
              'compute FIC, and record assay method, pathogen strain, and observed '
              'mechanism of action.'},
        {'q': 'Can I bulk-download the data?',
         'a': 'Yes. The Download page exports any filtered subset (or the full database) '
              'as CSV. Programmatic access is also available via the REST API at '
              '/api/v1/experiments/.'},
        {'q': 'How do I cite PhytoSynergyDB?',
         'a': 'Cite the database URL plus the original publication DOI for each '
              'experiment you reference. A formal citation entry will be issued on the '
              'first major release; until then, please cite by URL and access date.'},
        {'q': 'How much data can the API return at once?',
         'a': 'The public API is paginated: each request returns up to 500 records '
              '(default 100) via the limit and offset parameters. Iterate with '
              'offset to retrieve the full result set. For large computational '
              'pipelines, the full database is also available as a single CSV from '
              'the Download page.'},
        {'q': 'Is the database open for public submissions?',
         'a': 'No. PhytoSynergyDB is a closed, expert-curated resource maintained by '
              'the authoring team. Records are added in-house after manual extraction '
              'from peer-reviewed publications; public users have read-only access to '
              'the search, download and API features.'},
    ]

    context = {
        'synergy_entries_count': synergy_entries_count,
        'phytochemical_count': phytochemical_count,
        'antibiotic_count': antibiotic_count,
        'source_count': source_count,
        'eskape_pathogen_count': eskape_pathogen_count,
        'synergy_confirmed_count': synergy_confirmed_count,
        'synergy_share': synergy_share,
        'eskape_data': eskape_data,
        'recent_entries': recent_entries,
        'journals_row1': journals_row1,
        'journals_row2': journals_row2,
        'faq_data': faq_data,
    }
    return render(request, 'synergy_data/home.html', context)


# ==============================================================================
# ABOUT PAGE
# ==============================================================================

def about_page(request):
    total_experiments = SynergyExperiment.objects.count()
    total_phytochemicals = Phytochemical.objects.count()
    total_antibiotics = Antibiotic.objects.count()
    total_pathogens = Pathogen.objects.count()
    total_sources = Source.objects.count()
    total_synergy = SynergyExperiment.objects.filter(interpretation='Synergy').count()
    last_curation_date = timezone.now().strftime('%B %Y')

    about_faqs = [
        {'q': 'How is the FIC index calculated when MIC units differ between studies?',
         'a': 'All MIC values are normalised to a single declared unit (default µg/mL) '
              'during curation. The FIC index is unit-free because it is a ratio, but '
              'we record the original units alongside each value for traceability.'},
        {'q': 'Is the database open for public submissions?',
         'a': 'No. PhytoSynergyDB is a closed, expert-curated resource maintained by '
              'the authoring team. Records are added in-house after manual extraction '
              'from peer-reviewed publications; public users have read-only access.'},
        {'q': 'How are duplicate experiments resolved?',
         'a': 'A composite key of (phytochemical, antibiotic, pathogen, source) is '
              'enforced. If a row matches an existing experiment on all four fields '
              'during bulk import it is reported and skipped, not overwritten.'},
        {'q': 'What licence governs reuse of the data?',
         'a': 'The curated dataset is released under Creative Commons Attribution 4.0 '
              '(CC-BY-4.0). The source code is open under the MIT licence. Please '
              'cite PhytoSynergyDB and the underlying chemistry sources where '
              'descriptors are reused.'},
        {'q': 'How fresh is the database?',
         'a': 'New literature is curated on a rolling basis. Each record links to its '
              'primary publication via DOI, and bulk re-enrichment against PubChem '
              'and ClassyFire is run after every import cycle.'},
    ]

    context = {
        'total_experiments': total_experiments,
        'total_phytochemicals': total_phytochemicals,
        'total_antibiotics': total_antibiotics,
        'total_pathogens': total_pathogens,
        'total_sources': total_sources,
        'total_synergy': total_synergy,
        'last_curation_date': last_curation_date,
        'about_faqs': about_faqs,
    }
    return render(request, 'synergy_data/about.html', context)


# ==============================================================================
# DATABASE SEARCH PAGE
# ==============================================================================

def database_search_page(request):
    results = SynergyExperiment.objects.select_related(
        'phytochemical', 'antibiotic', 'pathogen', 'source'
    ).all()

    results = _apply_search_filters(results, request)
    results = results.order_by('-id')
    total_results = results.count()

    paginator = Paginator(results, 25)
    page_param = request.GET.get('page')
    try:
        page_obj = paginator.page(page_param)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Build a querystring with all current filters except `page`, so pagination
    # links can preserve filters.
    qd = request.GET.copy()
    qd.pop('page', None)
    preserved_qs = qd.urlencode()

    # Get distinct chemical classes for filter dropdown
    chemical_classes = (
        Phytochemical.objects.values_list('chemical_class', flat=True)
        .exclude(chemical_class__isnull=True)
        .exclude(chemical_class__exact='')
        .distinct()
        .order_by('chemical_class')
    )

    context = {
        'results': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_results': total_results,
        'preserved_qs': preserved_qs,
        'pathogens': Pathogen.objects.order_by('genus', 'species').all(),
        'antibiotics': Antibiotic.objects.order_by('antibiotic_name').all(),
        'mechanisms': SynergyExperiment.objects.values_list(
            'moa_observed', flat=True
        ).distinct().exclude(moa_observed__isnull=True).exclude(moa_observed__exact=''),
        'chemical_classes': chemical_classes,
        'search_query': request.GET.get('query', ''),
        'selected_pathogen': request.GET.get('pathogen'),
        'selected_antibiotic': request.GET.get('antibiotic'),
        'selected_mechanism': request.GET.get('mechanism'),
        'selected_interpretation': request.GET.get('interpretation'),
        'selected_eskape': request.GET.get('eskape'),
        'selected_chemical_class': request.GET.get('chemical_class'),
    }
    return render(request, 'synergy_data/database_search.html', context)


# ==============================================================================
# CHEMICAL SIMILARITY SEARCH (public)
# ==============================================================================

# A few well-known phytochemical SMILES offered as one-click examples on the
# similarity search page.
SIMILARITY_EXAMPLES = [
    {'name': 'Quercetin', 'smiles': 'O=c1c(O)c(-c2ccc(O)c(O)c2)oc2cc(O)cc(O)c12'},
    {'name': 'Berberine', 'smiles': 'COc1ccc2cc3[n+](cc2c1OC)CCc1cc2c(cc1-3)OCO2'},
    {'name': 'Curcumin', 'smiles': 'O=C(/C=C/c1ccc(O)c(OC)c1)CC(=O)/C=C/c1ccc(O)c(OC)c1'},
    {'name': 'Thymol', 'smiles': 'Cc1ccc(C(C)C)cc1O'},
]


def _searchable_phyto_qs():
    """Phytochemicals that carry a SMILES (the searchable universe)."""
    return (
        Phytochemical.objects
        .exclude(canonical_smiles__isnull=True)
        .exclude(canonical_smiles__exact='')
    )


def similarity_search_page(request):
    """Rank database phytochemicals by Tanimoto similarity (ECFP4) to a query SMILES."""
    query_smiles = (request.GET.get('smiles') or '').strip()

    try:
        limit = min(max(int(request.GET.get('limit', 25)), 1), 100)
    except (ValueError, TypeError):
        limit = 25
    try:
        threshold = min(max(float(request.GET.get('threshold', 0)), 0.0), 1.0)
    except (ValueError, TypeError):
        threshold = 0.0

    results = []
    error = None
    searched = bool(query_smiles)

    if searched:
        try:
            ranked = similarity.search_similar(
                query_smiles, limit=limit, threshold=threshold
            )
        except ValueError as exc:
            error = str(exc)
        except RuntimeError as exc:
            error = str(exc)
        else:
            for r in ranked:
                phyto = r['phytochemical']
                exps = SynergyExperiment.objects.filter(phytochemical=phyto)
                results.append({
                    'phytochemical': phyto,
                    'similarity': r['similarity'],
                    'percent': round(r['similarity'] * 100, 1),
                    'experiment_count': exps.count(),
                    'synergy_count': exps.filter(interpretation='Synergy').count(),
                })

    context = {
        'query_smiles': query_smiles,
        'results': results,
        'error': error,
        'searched': searched,
        'threshold': threshold,
        'limit': limit,
        'total_searchable': _searchable_phyto_qs().count(),
        'examples': SIMILARITY_EXAMPLES,
    }
    return render(request, 'synergy_data/similarity_search.html', context)


# ==============================================================================
# DATA ENTRY PAGE (login-protected)
# ==============================================================================

@login_required
def data_entry_view(request):
    if request.method == 'POST':
        form = SynergyEntryForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            # 1. Resolve Source (get_or_create by DOI, update metadata)
            doi = cd['source_doi'].strip()

            # Known-paper warning: if this DOI/title already exists and the
            # curator hasn't acknowledged it yet, re-render the form with a
            # yellow banner showing how many experiments are already on file.
            if request.POST.get('confirm_known_paper') != '1':
                existing_src = _lookup_existing_paper(
                    doi=doi,
                    pmid=cd.get('pmid'),
                    title=cd.get('article_title'),
                )
                if existing_src is not None:
                    existing_count = SynergyExperiment.objects.filter(
                        source=existing_src
                    ).count()
                    return render(request, 'synergy_data/data_entry.html', {
                        'form': form,
                        'known_paper_warning': {
                            'doi': existing_src.doi or '',
                            'title': existing_src.article_title or '',
                            'existing_count': existing_count,
                        },
                    })
            source, _ = Source.objects.get_or_create(doi=doi)
            # Update source metadata if provided (fills blanks)
            source_updated = False
            if cd.get('publication_year') and not source.publication_year:
                source.publication_year = cd['publication_year']
                source_updated = True
            if cd.get('article_title') and not source.article_title:
                source.article_title = cd['article_title'].strip()
                source_updated = True
            if cd.get('journal') and not source.journal:
                source.journal = cd['journal'].strip()
                source_updated = True
            if source_updated:
                source.save()

            # 2. Resolve Pathogen (parse name, set gram stain)
            pathogen = resolve_pathogen(
                cd['pathogen_full_name'], gram_stain=cd.get('gram_stain')
            )

            # 3. Resolve Phytochemical (case-insensitive)
            phytochemical = get_or_create_case_insensitive(
                Phytochemical, 'compound_name', cd['phytochemical_name'].strip()
            )

            # 3b. Auto-enrich phytochemical with PubChem + ClassyFire data
            enrichment_status = enrich_phytochemical(phytochemical)

            # 3b-ii. Compute the ECFP4 fingerprint for similarity search now that
            # a canonical SMILES is (likely) available from enrichment.
            similarity.update_fingerprint(phytochemical)

            # 3c. Link plant source (optional)
            link_plant_source(cd.get('plant_source'), phytochemical)

            # 4. Resolve Antibiotic (case-insensitive) + link class
            antibiotic = resolve_antibiotic(
                cd['antibiotic_name'].strip(), class_name=cd.get('antibiotic_class')
            )

            # 5. Auto-calculate FIC if not provided
            fic_index = cd.get('fic_index')
            if fic_index is None:
                fic_index = auto_calculate_fic(
                    cd.get('mic_phyto_alone'),
                    cd.get('mic_abx_alone'),
                    cd.get('mic_phyto_in_combo'),
                    cd.get('mic_abx_in_combo'),
                )

            # 6. Auto-interpret if not provided
            interpretation = cd.get('interpretation') or auto_interpret_fic(fic_index)

            # 7. Create the experiment record
            experiment = SynergyExperiment.objects.create(
                phytochemical=phytochemical,
                antibiotic=antibiotic,
                pathogen=pathogen,
                source=source,
                mic_phyto_alone=cd.get('mic_phyto_alone'),
                mic_abx_alone=cd.get('mic_abx_alone'),
                mic_phyto_in_combo=cd.get('mic_phyto_in_combo'),
                mic_abx_in_combo=cd.get('mic_abx_in_combo'),
                mic_units=cd.get('mic_units') or '\u00b5g/mL',
                fic_index=fic_index,
                interpretation=interpretation,
                assay_method=cd.get('assay_method') or 'checkerboard',
                moa_observed=cd.get('moa_observed', ''),
                notes=cd.get('notes', ''),
            )

            fic_display = f"{fic_index:.4f}" if fic_index else "N/A"
            enrichment_parts = []
            if enrichment_status.get("pubchem"):
                enrichment_parts.append("\u2713 PubChem enriched")
            if enrichment_status.get("classyfire"):
                enrichment_parts.append("\u2713 ClassyFire classified")
            enrichment_msg = (" | " + " | ".join(enrichment_parts)) if enrichment_parts else ""
            messages.success(
                request,
                f"Entry saved: {experiment}. FIC={fic_display}, "
                f"Interpretation={interpretation or 'N/A'}{enrichment_msg}"
            )

            # "Save & Add Another" pre-fills DOI, year, journal and MIC units
            if 'save_and_another' in request.POST:
                form = SynergyEntryForm(initial={
                    'source_doi': doi,
                    'publication_year': cd.get('publication_year'),
                    'article_title': cd.get('article_title'),
                    'journal': cd.get('journal'),
                    'mic_units': cd.get('mic_units') or '\u00b5g/mL',
                })
            else:
                return redirect('database_search')
    else:
        form = SynergyEntryForm()

    return render(request, 'synergy_data/data_entry.html', {'form': form})


# ==============================================================================
# EDIT ENTRY PAGE (login-protected)
# ==============================================================================

@login_required
def edit_entry_view(request, pk):
    """Edit an existing synergy experiment record."""
    from django.shortcuts import get_object_or_404

    experiment = get_object_or_404(
        SynergyExperiment.objects.select_related(
            'phytochemical', 'antibiotic', 'pathogen', 'source'
        ),
        pk=pk,
    )

    if request.method == 'POST':
        form = SynergyEntryForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            # 1. Resolve Source
            doi = cd['source_doi'].strip()
            source, _ = Source.objects.get_or_create(doi=doi)
            source_updated = False
            if cd.get('publication_year'):
                source.publication_year = cd['publication_year']
                source_updated = True
            if cd.get('article_title'):
                source.article_title = cd['article_title'].strip()
                source_updated = True
            if cd.get('journal'):
                source.journal = cd['journal'].strip()
                source_updated = True
            if source_updated:
                source.save()

            # 2. Resolve Pathogen
            pathogen = resolve_pathogen(
                cd['pathogen_full_name'], gram_stain=cd.get('gram_stain')
            )

            # 3. Resolve Phytochemical
            phytochemical = get_or_create_case_insensitive(
                Phytochemical, 'compound_name', cd['phytochemical_name'].strip()
            )

            # 3b. Auto-enrich
            enrichment_status = enrich_phytochemical(phytochemical)

            # 3b-ii. Refresh the ECFP4 fingerprint for similarity search.
            similarity.update_fingerprint(phytochemical)

            # 3c. Link plant source (optional)
            link_plant_source(cd.get('plant_source'), phytochemical)

            # 4. Resolve Antibiotic + link class
            antibiotic = resolve_antibiotic(
                cd['antibiotic_name'].strip(), class_name=cd.get('antibiotic_class')
            )

            # 5. Auto-calculate FIC if not provided
            fic_index = cd.get('fic_index')
            if fic_index is None:
                fic_index = auto_calculate_fic(
                    cd.get('mic_phyto_alone'),
                    cd.get('mic_abx_alone'),
                    cd.get('mic_phyto_in_combo'),
                    cd.get('mic_abx_in_combo'),
                )

            # 6. Auto-interpret
            interpretation = cd.get('interpretation') or auto_interpret_fic(fic_index)

            # 7. Update the experiment record (not create!)
            experiment.phytochemical = phytochemical
            experiment.antibiotic = antibiotic
            experiment.pathogen = pathogen
            experiment.source = source
            experiment.mic_phyto_alone = cd.get('mic_phyto_alone')
            experiment.mic_abx_alone = cd.get('mic_abx_alone')
            experiment.mic_phyto_in_combo = cd.get('mic_phyto_in_combo')
            experiment.mic_abx_in_combo = cd.get('mic_abx_in_combo')
            experiment.mic_units = cd.get('mic_units') or '\u00b5g/mL'
            experiment.fic_index = fic_index
            experiment.interpretation = interpretation
            experiment.assay_method = cd.get('assay_method') or 'checkerboard'
            experiment.moa_observed = cd.get('moa_observed', '')
            experiment.notes = cd.get('notes', '')
            experiment.save()

            fic_display = f"{fic_index:.4f}" if fic_index else "N/A"
            enrichment_parts = []
            if enrichment_status.get("pubchem"):
                enrichment_parts.append("\u2713 PubChem enriched")
            if enrichment_status.get("classyfire"):
                enrichment_parts.append("\u2713 ClassyFire classified")
            enrichment_msg = (" | " + " | ".join(enrichment_parts)) if enrichment_parts else ""
            messages.success(
                request,
                f"Entry updated: {experiment}. FIC={fic_display}, "
                f"Interpretation={interpretation or 'N/A'}{enrichment_msg}"
            )
            return redirect('database_search')
    else:
        # Pre-populate form with existing data
        pathogen_str = f"{experiment.pathogen.genus} {experiment.pathogen.species}"
        if experiment.pathogen.strain:
            pathogen_str += f" {experiment.pathogen.strain}"

        existing_plant = experiment.phytochemical.source_plants.first()
        form = SynergyEntryForm(initial={
            'source_doi': experiment.source.doi or '',
            'publication_year': experiment.source.publication_year,
            'article_title': experiment.source.article_title or '',
            'journal': experiment.source.journal or '',
            'pathogen_full_name': pathogen_str,
            'gram_stain': experiment.pathogen.gram_stain or '',
            'phytochemical_name': experiment.phytochemical.compound_name,
            'plant_source': existing_plant.scientific_name if existing_plant else '',
            'antibiotic_name': experiment.antibiotic.antibiotic_name,
            'antibiotic_class': (
                experiment.antibiotic.antibiotic_class.class_name
                if experiment.antibiotic.antibiotic_class else ''
            ),
            'assay_method': experiment.assay_method or 'checkerboard',
            'mic_phyto_alone': experiment.mic_phyto_alone,
            'mic_abx_alone': experiment.mic_abx_alone,
            'mic_phyto_in_combo': experiment.mic_phyto_in_combo,
            'mic_abx_in_combo': experiment.mic_abx_in_combo,
            'mic_units': experiment.mic_units or '\u00b5g/mL',
            'fic_index': experiment.fic_index,
            'interpretation': experiment.interpretation or '',
            'moa_observed': experiment.moa_observed or '',
            'notes': experiment.notes or '',
        })

    return render(request, 'synergy_data/data_entry.html', {
        'form': form,
        'editing': True,
        'experiment_id': pk,
    })


# ==============================================================================
# BULK CSV / XLSX IMPORT (login-protected)
# ==============================================================================

# Strings students commonly type for "empty" cells - all treated as NULL.
_NULL_STRINGS = {
    'null', 'n/a', 'na', 'none', '-', '--', '---', '', 'not reported',
    'not available', 'nr', 'nd', 'not determined', 'unknown', '?',
}


def _clean_value(value, field_type='text'):
    """Clean a single cell value from student-entered data.

    Handles:
    - literal "null" strings -> None
    - corrupted micro signs ("Âµg/mL" -> "µg/mL")
    - tabs, NBSP, carriage returns
    - decimal ranges ("32-64" -> 64.0, upper bound)
    - inequality prefixes (">256" -> 256.0)
    - year strings with trailing ".0" from Excel numeric cells
    """
    if value is None:
        return None

    s = str(value).strip()

    # Fast path: literal None text or empty
    if s.lower() in _NULL_STRINGS:
        return None

    # Encoding repair - order matters, the longer sequences go first.
    s = (
        s.replace('Âµ', 'µ')   # CP1252 double-decode of U+00B5
         .replace('Î¼', 'µ')   # CP1252 double-decode of U+03BC (Greek mu)
         .replace('\t', '')
         .replace('\r', '')
         .replace('\xa0', ' ')
    ).strip()

    if s == '' or s.lower() in _NULL_STRINGS:
        return None

    if field_type == 'decimal':
        # Range like "32-64" - take the upper bound (not a leading-negative).
        if '-' in s and not s.startswith('-'):
            parts = [p.strip() for p in s.split('-') if p.strip()]
            try:
                return Decimal(str(max(float(p) for p in parts)))
            except (ValueError, InvalidOperation):
                return None

        # Strip inequality markers - we only store a point estimate.
        for ch in ('>', '<', '≥', '≤', '=', '~'):
            s = s.replace(ch, '')
        s = s.strip()

        try:
            return Decimal(s)
        except (InvalidOperation, ValueError, TypeError):
            return None

    if field_type == 'integer':
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None

    return s


def _safe_decimal(value):
    """Back-compat shim used elsewhere in this module."""
    return _clean_value(value, field_type='decimal')


# Controlled vocabulary for SynergyExperiment.assay_method (must match models.py).
VALID_ASSAY_METHODS = {
    'checkerboard', 'time_kill', 'disk_diffusion', 'broth_microdilution', 'other',
}
# Keyword -> canonical code, checked in order against free-text descriptions.
_ASSAY_KEYWORD_MAP = [
    ('checkerboard', 'checkerboard'),
    ('time', 'time_kill'),
    ('kill', 'time_kill'),
    ('disk', 'disk_diffusion'),
    ('disc', 'disk_diffusion'),
    ('diffusion', 'disk_diffusion'),
    ('broth', 'broth_microdilution'),
    ('microdilution', 'broth_microdilution'),
    ('microbroth', 'broth_microdilution'),
]


def normalize_assay_method(raw_value, notes=''):
    """Map a free-text or oversized assay_method into a valid controlled code.

    The DB column is varchar(50) with a fixed choice list. Imports sometimes
    carry a long methodology sentence (e.g. "Checkerboard microbroth dilution
    (8x8 matrix, MH broth ...)") which overflows the column. This collapses such
    text to a canonical code by keyword match (falling back to 'other') and, when
    the original text was discarded, preserves it by prepending to ``notes``.

    Returns (code, notes) - both ready to assign directly to the model.
    """
    value = (str(raw_value).strip() if raw_value is not None else '')
    notes = notes or ''
    if not value:
        return 'checkerboard', notes

    lowered = value.lower()
    if lowered in VALID_ASSAY_METHODS:
        return lowered, notes

    code = 'other'
    for keyword, mapped in _ASSAY_KEYWORD_MAP:
        if keyword in lowered:
            code = mapped
            break

    # Preserve the original description so curated detail is not lost.
    prefix = f"Assay method: {value}"
    if prefix not in notes:
        notes = f"{prefix}\n{notes}".strip() if notes else prefix
    return code, notes


def _parse_upload_to_rows(uploaded_file):
    """Parse either a CSV or XLSX upload into a list of dicts keyed by
    canonical column names. Returns (rows, total_row_count)."""
    name = uploaded_file.name.lower()

    raw_headers = []
    raw_rows = []

    if name.endswith('.xlsx'):
        import openpyxl  # local import keeps startup light
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        try:
            raw_headers = list(next(iterator))
        except StopIteration:
            return [], 0
        for row in iterator:
            raw_rows.append(list(row))
    else:
        uploaded_file.seek(0)
        text = uploaded_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(text))
        try:
            raw_headers = next(reader)
        except StopIteration:
            return [], 0
        for row in reader:
            raw_rows.append(row)

    # Map each raw header to its canonical internal name (empty string = ignore).
    canonical_headers = [_canonical_header(h) for h in raw_headers]

    cleaned_rows = []
    for raw in raw_rows:
        # Pad short rows, trim long rows to header length.
        if len(raw) < len(canonical_headers):
            raw = list(raw) + [None] * (len(canonical_headers) - len(raw))
        row_dict = {}
        for key, val in zip(canonical_headers, raw):
            if not key:
                continue
            # Last-column-wins if a file has duplicate headers after aliasing.
            row_dict[key] = val

        # Drop rows where every cell is empty / None.
        if not any(
            (v is not None and str(v).strip() != '')
            for v in row_dict.values()
        ):
            continue
        cleaned_rows.append(row_dict)

    return cleaned_rows, len(cleaned_rows)


# Columns that must be cleaned as decimals when staging a row.
_DECIMAL_FIELDS = (
    'mic_phyto_alone', 'mic_abx_alone',
    'mic_phyto_in_combo', 'mic_abx_in_combo',
    'fic_index',
)
_INTEGER_FIELDS = ('publication_year', 'pmid')


def _stage_row(row_num, raw_row):
    """Apply cleaning + validation to a single raw row dict and classify it.

    Returns a dict with keys: status ('valid' | 'warning' | 'error'),
    row_num, data (cleaned dict), reason (str, for errors/warnings),
    fic (Decimal|None), interpretation (str|None).
    """
    clean = {}
    for key, val in raw_row.items():
        if key in _DECIMAL_FIELDS:
            clean[key] = _clean_value(val, 'decimal')
        elif key in _INTEGER_FIELDS:
            clean[key] = _clean_value(val, 'integer')
        else:
            clean[key] = _clean_value(val, 'text')

    # Compose the pathogen full name from split genus/species/strain columns
    # when they are provided (preferred). Falls back to the legacy single
    # pathogen_full_name column for older sheets.
    genus = clean.get('pathogen_genus')
    if genus and not clean.get('pathogen_full_name'):
        parts = [genus, clean.get('pathogen_species'), clean.get('pathogen_strain')]
        clean['pathogen_full_name'] = ' '.join(p for p in parts if p).strip()

    # Required fields
    required = ('source_doi', 'pathogen_full_name', 'phytochemical_name', 'antibiotic_name')
    missing_required = [f for f in required if not clean.get(f)]
    if missing_required:
        label = {'pathogen_full_name': 'pathogen (genus + species)'}
        return {
            'status': 'error',
            'row_num': row_num,
            'data': clean,
            'reason': "Missing required fields: "
                      + ', '.join(label.get(f, f) for f in missing_required),
            'fic': None,
            'interpretation': None,
        }

    # Reject abbreviated genera (e.g. "S." for Staphylococcus) - the cause of
    # the malformed-genus data. The genus is the first token of the full name.
    first_token = clean['pathogen_full_name'].split()[0]
    if len(first_token) <= 2 and first_token.endswith('.'):
        return {
            'status': 'error',
            'row_num': row_num,
            'data': clean,
            'reason': f"Genus '{first_token}' is abbreviated - use the full genus "
                      f"name (e.g. 'Staphylococcus', not 'S.').",
            'fic': None,
            'interpretation': None,
        }

    mic_vals = [
        clean.get('mic_phyto_alone'),
        clean.get('mic_abx_alone'),
        clean.get('mic_phyto_in_combo'),
        clean.get('mic_abx_in_combo'),
    ]
    has_all_mic = all(v is not None for v in mic_vals)
    fic_val = clean.get('fic_index')

    if not has_all_mic and fic_val is None:
        return {
            'status': 'error',
            'row_num': row_num,
            'data': clean,
            'reason': (
                "Must have all 4 MIC values or a FIC index. "
                "Qualitative-only data (disk diffusion, zone diameters) not accepted."
            ),
            'fic': None,
            'interpretation': None,
        }

    # Auto-calc FIC if all 4 MICs are present and FIC is missing.
    if fic_val is None and has_all_mic:
        fic_val = auto_calculate_fic(*mic_vals)

    # Normalise interpretation free-text onto the controlled vocabulary.
    interp_raw = clean.get('interpretation')
    interpretation_map = {
        'synergy': 'Synergy', 'synergistic': 'Synergy', 'syn': 'Synergy',
        'additive': 'Additive', 'add': 'Additive', 'partial synergy': 'Additive',
        'indifference': 'Indifference', 'indifferent': 'Indifference', 'ind': 'Indifference',
        'no interaction': 'Indifference',
        'antagonism': 'Antagonism', 'antagonistic': 'Antagonism', 'ant': 'Antagonism',
    }
    if interp_raw:
        interp = interpretation_map.get(interp_raw.lower().strip(), interp_raw)
    else:
        interp = None
    if interp not in ('Synergy', 'Additive', 'Indifference', 'Antagonism'):
        interp = auto_interpret_fic(fic_val)

    # Warning if optional-but-useful data is missing (MIC values when only FIC
    # provided, or missing interpretation/MOA). Row is still importable.
    warnings = []
    if not has_all_mic:
        warnings.append("Only FIC provided - some MIC values missing")
    if not clean.get('moa_observed'):
        warnings.append("No mechanism of action")
    if not clean.get('publication_year'):
        warnings.append("No publication year")

    status = 'warning' if warnings else 'valid'
    return {
        'status': status,
        'row_num': row_num,
        'data': clean,
        'reason': '; '.join(warnings) if warnings else '',
        'fic': fic_val,
        'interpretation': interp,
    }


def _row_already_imported(clean):
    """Read-only check: is this row already in the database?

    Mirrors the confirm-step duplicate key (phytochemical + antibiotic +
    pathogen + source) WITHOUT creating any FK rows, so it is safe to call
    during the preview pass. Returns False as soon as any FK is missing - a
    row referencing a not-yet-seen paper/compound/etc. cannot be a duplicate.
    """
    source = Source.objects.filter(doi=clean.get('source_doi')).first()
    if source is None:
        return False

    name = (clean.get('phytochemical_name') or '').strip()
    phyto = Phytochemical.objects.filter(compound_name__iexact=name).first()
    if phyto is None:
        return False

    abx_name = (clean.get('antibiotic_name') or '').strip()
    abx = Antibiotic.objects.filter(antibiotic_name__iexact=abx_name).first()
    if abx is None:
        return False

    genus, species, strain = parse_pathogen_name(clean.get('pathogen_full_name') or '')
    pathogen = Pathogen.objects.filter(
        genus=genus, species=species, strain=strain
    ).first()
    if pathogen is None:
        return False

    return SynergyExperiment.objects.filter(
        phytochemical=phyto, antibiotic=abx, pathogen=pathogen, source=source,
    ).exists()


@login_required
def bulk_import_template(request):
    """Download a blank XLSX template for bulk import, pre-filled with one
    example row so students can see the expected format and units."""
    try:
        import openpyxl
    except ImportError:
        # Graceful fallback to CSV if openpyxl isn't installed yet.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            'attachment; filename="PhytoSynergyDB_import_template.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(BULK_CSV_COLUMNS)
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PhytoSynergyDB Data"

    for col_idx, header in enumerate(BULK_CSV_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    example = {
        'source_doi': '10.1016/j.phymed.2023.154789',
        'pmid': '',
        'publication_year': 2023,
        'article_title': 'Synergistic activity of berberine with ciprofloxacin against S. aureus',
        'journal': 'Phytomedicine',
        'pathogen_genus': 'Staphylococcus',
        'pathogen_species': 'aureus',
        'pathogen_strain': 'ATCC 25923',
        'gram_stain': '',
        'phytochemical_name': 'Berberine',
        'plant_source': 'Berberis vulgaris',
        'antibiotic_name': 'Ciprofloxacin',
        'antibiotic_class': 'Fluoroquinolone',
        'mic_phyto_alone': 64,
        'mic_abx_alone': 0.5,
        'mic_phyto_in_combo': 16,
        'mic_abx_in_combo': 0.125,
        'mic_units': 'µg/mL',
        'fic_index': 0.5,
        'interpretation': 'Synergy',
        'assay_method': 'checkerboard',
        'moa_observed': 'Efflux pump inhibition',
        'notes': '',
    }
    for col_idx, header in enumerate(BULK_CSV_COLUMNS, start=1):
        ws.cell(row=2, column=col_idx, value=example.get(header, ''))

    # Dropdown (data-validation) lists to cut down typos.
    from openpyxl.worksheet.datavalidation import DataValidation
    dropdowns = {
        'gram_stain': '"Gram-positive,Gram-negative"',
        'interpretation': '"Synergy,Additive,Indifference,Antagonism"',
        'assay_method': '"checkerboard,time_kill,disk_diffusion,broth_microdilution,other"',
    }
    for field, formula in dropdowns.items():
        if field not in BULK_CSV_COLUMNS:
            continue
        col_letter = openpyxl.utils.get_column_letter(BULK_CSV_COLUMNS.index(field) + 1)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}1000")

    # Reasonable default widths
    for col_idx in range(1, len(BULK_CSV_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="PhytoSynergyDB_import_template.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def bulk_import_view(request):
    """Bulk CSV/XLSX import with strict validation.

    Every row must have all 4 MICs or a FIC index. Rows missing only optional
    fields (MOA, year) are flagged as warnings but still imported on confirm.
    Duplicate experiments (same phyto+abx+pathogen+source) are skipped.
    """
    context = {'form': BulkCSVUploadForm()}

    # Surface the outcome of a just-completed import as one summary card.
    context['import_result'] = request.session.pop('bulk_import_result', None)

    if request.method == 'POST':
        action = request.POST.get('action', 'validate')

        # ── CONFIRM ACTION: save previously validated rows ──
        if action == 'confirm':
            imported = 0
            skipped_duplicate = 0
            skipped_invalid = 0
            errors = []

            confirm_json = request.POST.get('import_data_json', '[]')
            try:
                rows_to_save = json.loads(confirm_json)
            except json.JSONDecodeError:
                messages.error(request, "Could not parse confirmed data.")
                return redirect('bulk_import')

            for row_wrapper in rows_to_save:
                if row_wrapper.get('status') == 'error':
                    skipped_invalid += 1
                    continue

                row = row_wrapper.get('data', {})
                row_num = row_wrapper.get('row_num', 0)
                try:
                    # Each row is its own atomic unit: either every
                    # get_or_create + the experiment row commit together, or
                    # nothing for this row commits. Prevents orphan FK records
                    # if the worker is killed mid-loop.
                    with transaction.atomic():
                        # --- Source (get_or_create by DOI) ---
                        source, _ = Source.objects.get_or_create(
                            doi=row.get('source_doi'),
                            defaults={
                                'pmid': row.get('pmid'),
                                'publication_year': row.get('publication_year'),
                                'article_title': row.get('article_title') or None,
                                'journal': row.get('journal') or None,
                            },
                        )

                        # --- Pathogen (parse + get_or_create, derive gram) ---
                        pathogen = resolve_pathogen(
                            row['pathogen_full_name'],
                            gram_stain=row.get('gram_stain'),
                        )

                        # --- Phytochemical (case-insensitive) ---
                        # NOTE: PubChem/ClassyFire enrichment is intentionally
                        # NOT called here. Doing 6+ network calls per row inside
                        # a request blocks the gunicorn worker (~46 s/row worst
                        # case) and freezes the whole site. Run
                        #   python manage.py enrich_phytochemicals
                        # after the import to backfill chemistry data offline.
                        phytochemical = get_or_create_case_insensitive(
                            Phytochemical, 'compound_name', row['phytochemical_name']
                        )

                        # --- Plant source (optional M2M) ---
                        link_plant_source(row.get('plant_source'), phytochemical)

                        # --- Antibiotic (case-insensitive) + class ---
                        antibiotic = resolve_antibiotic(
                            row['antibiotic_name'],
                            class_name=row.get('antibiotic_class'),
                        )

                        # --- Duplicate check ---
                        if SynergyExperiment.objects.filter(
                            phytochemical=phytochemical,
                            antibiotic=antibiotic,
                            pathogen=pathogen,
                            source=source,
                        ).exists():
                            skipped_duplicate += 1
                            continue

                        # --- MIC values (re-clean in case the JSON round-trip
                        #     turned them into strings) ---
                        mic_phyto_alone = _safe_decimal(row.get('mic_phyto_alone'))
                        mic_abx_alone = _safe_decimal(row.get('mic_abx_alone'))
                        mic_phyto_in_combo = _safe_decimal(row.get('mic_phyto_in_combo'))
                        mic_abx_in_combo = _safe_decimal(row.get('mic_abx_in_combo'))
                        mic_units = (row.get('mic_units') or 'µg/mL').strip() or 'µg/mL'

                        fic_index = _safe_decimal(row.get('fic_index'))
                        if fic_index is None:
                            fic_index = auto_calculate_fic(
                                mic_phyto_alone, mic_abx_alone,
                                mic_phyto_in_combo, mic_abx_in_combo,
                            )
                        interpretation = (row.get('interpretation') or '').strip()
                        if interpretation not in ['Synergy', 'Additive', 'Indifference', 'Antagonism']:
                            interpretation = auto_interpret_fic(fic_index)

                        assay_method, notes = normalize_assay_method(
                            row.get('assay_method'), row.get('notes') or '',
                        )

                        SynergyExperiment.objects.create(
                            phytochemical=phytochemical,
                            antibiotic=antibiotic,
                            pathogen=pathogen,
                            source=source,
                            mic_phyto_alone=mic_phyto_alone,
                            mic_abx_alone=mic_abx_alone,
                            mic_phyto_in_combo=mic_phyto_in_combo,
                            mic_abx_in_combo=mic_abx_in_combo,
                            mic_units=mic_units,
                            fic_index=fic_index,
                            interpretation=interpretation,
                            assay_method=assay_method,
                            moa_observed=row.get('moa_observed') or '',
                            notes=notes,
                        )
                        imported += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

            # Stash a single structured result for a clean summary card on the
            # next GET, instead of emitting one full-width alert per row error.
            request.session['bulk_import_result'] = {
                'imported': imported,
                'skipped_duplicate': skipped_duplicate,
                'skipped_invalid': skipped_invalid,
                'errors': errors,
            }
            return redirect('bulk_import')

        # ── VALIDATE ACTION: parse file and show preview ──
        form = BulkCSVUploadForm(request.POST, request.FILES)
        context['form'] = form

        if not form.is_valid():
            return render(request, 'synergy_data/bulk_import.html', context)

        uploaded_file = form.cleaned_data['csv_file']
        try:
            raw_rows, _total = _parse_upload_to_rows(uploaded_file)
        except Exception as e:
            messages.error(request, f"Could not parse uploaded file: {e}")
            return render(request, 'synergy_data/bulk_import.html', context)

        staged = []
        for offset, raw_row in enumerate(raw_rows, start=2):  # row 1 is header
            staged.append(_stage_row(offset, raw_row))

        # Flag rows that already exist in the DB so the preview reflects what
        # Confirm will actually do (those rows are skipped, not re-imported).
        for s in staged:
            if s['status'] in ('valid', 'warning') and _row_already_imported(s['data']):
                s['status'] = 'duplicate'
                s['reason'] = 'Already in the database - will be skipped'

        valid_count = sum(1 for s in staged if s['status'] == 'valid')
        warning_count = sum(1 for s in staged if s['status'] == 'warning')
        duplicate_count = sum(1 for s in staged if s['status'] == 'duplicate')
        error_count = sum(1 for s in staged if s['status'] == 'error')

        # Build the template-friendly view + a JSON blob for the confirm step.
        preview_rows = []
        import_payload = []
        for s in staged:
            preview_rows.append({
                'row_num': s['row_num'],
                'status': s['status'],
                'reason': s['reason'],
                'fic': s['fic'],
                'interpretation': s['interpretation'],
                'data': s['data'],
            })
            # Only importable rows survive the JSON round-trip. Errors and
            # already-in-DB duplicates are shown in preview but never imported.
            if s['status'] in ('valid', 'warning'):
                payload_data = {k: ('' if v is None else str(v)) for k, v in s['data'].items()}
                # Surface the computed FIC/interpretation so the confirm step
                # doesn't need to recalculate if the input FIC was blank.
                if s['fic'] is not None:
                    payload_data['fic_index'] = str(s['fic'])
                if s['interpretation']:
                    payload_data['interpretation'] = s['interpretation']
                import_payload.append({
                    'status': s['status'],
                    'row_num': s['row_num'],
                    'data': payload_data,
                })

        # Known-papers summary: which papers in this upload are already in the DB?
        # Group importable rows by paper key (normalized DOI, falling back to
        # PMID or article title). For each paper, count rows in this upload and
        # how many experiments already exist on file. Surfaced in the preview
        # template as a yellow banner so curators can sanity-check before
        # confirming.
        paper_buckets = {}
        for s in staged:
            if s['status'] == 'error':
                continue
            d = s['data']
            key = (
                _normalize_doi(d.get('source_doi'))
                or (d.get('pmid') or '').strip()
                or (d.get('article_title') or '').strip().lower()
            )
            if not key:
                continue
            paper_buckets.setdefault(key, {
                'doi': d.get('source_doi') or '',
                'title': d.get('article_title') or '',
                'rows_in_upload': 0,
            })['rows_in_upload'] += 1

        known_papers = []
        for key, info in paper_buckets.items():
            existing = _lookup_existing_paper(
                doi=info['doi'], title=info['title']
            )
            if existing is None:
                continue
            existing_count = SynergyExperiment.objects.filter(source=existing).count()
            known_papers.append({
                'doi': existing.doi or info['doi'],
                'title': existing.article_title or info['title'],
                'existing_count': existing_count,
                'rows_in_upload': info['rows_in_upload'],
            })

        context.update({
            'preview_rows': preview_rows,
            'valid_count': valid_count,
            'warning_count': warning_count,
            'duplicate_count': duplicate_count,
            'error_count': error_count,
            'importable_count': valid_count + warning_count,
            'total_count': len(staged),
            'import_data_json': json.dumps(import_payload, default=str),
            'known_papers': known_papers,
        })

    return render(request, 'synergy_data/bulk_import.html', context)


# ==============================================================================
# DATA DOWNLOAD (CSV Export)
# ==============================================================================

def download_data(request):
    """Export synergy experiment data as CSV."""
    if request.method == 'GET' and 'export' in request.GET:
        # Build queryset with filters
        results = SynergyExperiment.objects.select_related(
            'phytochemical', 'antibiotic', 'pathogen', 'source'
        ).all()
        results = _apply_search_filters(results, request)

        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="phytosynergydb_export.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Phytochemical', 'PubChem_CID', 'InChI_Key', 'SMILES',
            'Molecular_Formula', 'Molecular_Weight', 'XLogP',
            'HBond_Donors', 'HBond_Acceptors', 'TPSA', 'Rotatable_Bonds',
            'Lipinski_Pass', 'Chemical_Superclass', 'Chemical_Class', 'Chemical_Subclass',
            'Antibiotic', 'DrugBank_ID', 'Antibiotic_Class',
            'Pathogen_Genus', 'Pathogen_Species', 'Pathogen_Strain', 'Gram_Stain',
            'MIC_Phyto_Alone', 'MIC_Abx_Alone', 'MIC_Phyto_Combo', 'MIC_Abx_Combo', 'MIC_Units',
            'FIC_Index', 'Interpretation',
            'Mechanism_of_Action', 'Notes',
            'DOI', 'PMID', 'Journal', 'Publication_Year',
        ])

        for exp in results:
            lipinski = exp.phytochemical.passes_lipinski
            lipinski_str = ''
            if lipinski is True:
                lipinski_str = 'Yes'
            elif lipinski is False:
                lipinski_str = 'No'

            writer.writerow([
                exp.phytochemical.compound_name,
                exp.phytochemical.pubchem_cid or '',
                exp.phytochemical.inchi_key or '',
                exp.phytochemical.canonical_smiles or '',
                exp.phytochemical.molecular_formula or '',
                exp.phytochemical.molecular_weight or '',
                exp.phytochemical.xlogp if exp.phytochemical.xlogp is not None else '',
                exp.phytochemical.hbd if exp.phytochemical.hbd is not None else '',
                exp.phytochemical.hba if exp.phytochemical.hba is not None else '',
                exp.phytochemical.tpsa if exp.phytochemical.tpsa is not None else '',
                exp.phytochemical.rotatable_bonds if exp.phytochemical.rotatable_bonds is not None else '',
                lipinski_str,
                exp.phytochemical.chemical_superclass or '',
                exp.phytochemical.chemical_class or '',
                exp.phytochemical.chemical_subclass or '',
                exp.antibiotic.antibiotic_name,
                exp.antibiotic.drugbank_id or '',
                exp.antibiotic.antibiotic_class.class_name if exp.antibiotic.antibiotic_class else '',
                exp.pathogen.genus,
                exp.pathogen.species,
                exp.pathogen.strain or '',
                exp.pathogen.gram_stain or '',
                exp.mic_phyto_alone or '',
                exp.mic_abx_alone or '',
                exp.mic_phyto_in_combo or '',
                exp.mic_abx_in_combo or '',
                exp.mic_units,
                exp.fic_index or '',
                exp.interpretation or '',
                exp.moa_observed or '',
                exp.notes or '',
                exp.source.doi or '',
                exp.source.pmid or '',
                exp.source.journal or '',
                exp.source.publication_year or '',
            ])

        return response

    # Show download page
    total_experiments = SynergyExperiment.objects.count()
    total_synergies = SynergyExperiment.objects.filter(interpretation='Synergy').count()
    return render(request, 'synergy_data/download.html', {
        'total_experiments': total_experiments,
        'total_synergies': total_synergies,
    })


# ==============================================================================
# REST API ENDPOINTS
# ==============================================================================

def api_experiments(request):
    """JSON API endpoint for synergy experiments."""
    results = SynergyExperiment.objects.select_related(
        'phytochemical', 'antibiotic', 'pathogen', 'source'
    ).all()

    results = _apply_search_filters(results, request)

    # Pagination
    try:
        limit = min(int(request.GET.get('limit', 100)), 500)
    except (ValueError, TypeError):
        limit = 100
    try:
        offset = max(int(request.GET.get('offset', 0)), 0)
    except (ValueError, TypeError):
        offset = 0

    total_count = results.count()
    page_results = results[offset:offset + limit]

    data = []
    for exp in page_results:
        data.append({
            'id': exp.id,
            'phytochemical': {
                'name': exp.phytochemical.compound_name,
                'pubchem_cid': exp.phytochemical.pubchem_cid,
                'inchi_key': exp.phytochemical.inchi_key,
                'smiles': exp.phytochemical.canonical_smiles,
                'molecular_weight': str(exp.phytochemical.molecular_weight) if exp.phytochemical.molecular_weight else None,
                'molecular_formula': exp.phytochemical.molecular_formula,
                'xlogp': exp.phytochemical.xlogp,
                'hbd': exp.phytochemical.hbd,
                'hba': exp.phytochemical.hba,
                'tpsa': exp.phytochemical.tpsa,
                'rotatable_bonds': exp.phytochemical.rotatable_bonds,
                'passes_lipinski': exp.phytochemical.passes_lipinski,
                'chemical_superclass': exp.phytochemical.chemical_superclass,
                'chemical_class': exp.phytochemical.chemical_class,
                'chemical_subclass': exp.phytochemical.chemical_subclass,
            },
            'antibiotic': {
                'name': exp.antibiotic.antibiotic_name,
                'drugbank_id': exp.antibiotic.drugbank_id,
                'class': exp.antibiotic.antibiotic_class.class_name if exp.antibiotic.antibiotic_class else None,
            },
            'pathogen': {
                'genus': exp.pathogen.genus,
                'species': exp.pathogen.species,
                'strain': exp.pathogen.strain,
                'gram_stain': exp.pathogen.gram_stain,
            },
            'mic_phyto_alone': str(exp.mic_phyto_alone) if exp.mic_phyto_alone else None,
            'mic_abx_alone': str(exp.mic_abx_alone) if exp.mic_abx_alone else None,
            'mic_phyto_in_combo': str(exp.mic_phyto_in_combo) if exp.mic_phyto_in_combo else None,
            'mic_abx_in_combo': str(exp.mic_abx_in_combo) if exp.mic_abx_in_combo else None,
            'mic_units': exp.mic_units,
            'fic_index': str(exp.fic_index) if exp.fic_index else None,
            'interpretation': exp.interpretation,
            'mechanism_of_action': exp.moa_observed,
            'source': {
                'doi': exp.source.doi,
                'pmid': exp.source.pmid,
                'journal': exp.source.journal,
                'year': exp.source.publication_year,
                'title': exp.source.article_title,
            },
        })

    return JsonResponse({
        'count': total_count,
        'limit': limit,
        'offset': offset,
        'results': data,
    })


def api_statistics(request):
    """JSON API endpoint for database statistics."""
    synergy_count = SynergyExperiment.objects.filter(interpretation='Synergy').count()
    additive_count = SynergyExperiment.objects.filter(interpretation='Additive').count()
    indifference_count = SynergyExperiment.objects.filter(interpretation='Indifference').count()
    antagonism_count = SynergyExperiment.objects.filter(interpretation='Antagonism').count()

    eskape_stats = {}
    for genus in ['Enterococcus', 'Staphylococcus', 'Klebsiella', 'Acinetobacter', 'Pseudomonas', 'Enterobacter']:
        eskape_stats[genus] = SynergyExperiment.objects.filter(pathogen__genus=genus).count()

    return JsonResponse({
        'total_experiments': SynergyExperiment.objects.count(),
        'total_phytochemicals': Phytochemical.objects.count(),
        'total_antibiotics': Antibiotic.objects.count(),
        'total_sources': Source.objects.count(),
        'total_pathogens': Pathogen.objects.count(),
        'interpretations': {
            'synergy': synergy_count,
            'additive': additive_count,
            'indifference': indifference_count,
            'antagonism': antagonism_count,
        },
        'eskape_counts': eskape_stats,
    })


def api_similarity(request):
    """JSON API: rank database phytochemicals by Tanimoto similarity to a SMILES.

    Query params:
        smiles    (required) - the query structure as a SMILES string
        limit     (optional) - max hits to return (default 25, capped at 100)
        threshold (optional) - minimum Tanimoto similarity, 0.0 - 1.0 (default 0)
    """
    query_smiles = (request.GET.get('smiles') or '').strip()
    if not query_smiles:
        return JsonResponse(
            {'error': "Provide a 'smiles' query parameter, e.g. ?smiles=Cc1ccc(C(C)C)cc1O"},
            status=400,
        )

    try:
        limit = min(max(int(request.GET.get('limit', 25)), 1), 100)
    except (ValueError, TypeError):
        limit = 25
    try:
        threshold = min(max(float(request.GET.get('threshold', 0)), 0.0), 1.0)
    except (ValueError, TypeError):
        threshold = 0.0

    try:
        ranked = similarity.search_similar(query_smiles, limit=limit, threshold=threshold)
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    except RuntimeError as exc:
        return JsonResponse({'error': str(exc)}, status=503)

    data = []
    for r in ranked:
        phyto = r['phytochemical']
        exps = SynergyExperiment.objects.filter(phytochemical=phyto)
        data.append({
            'compound_name': phyto.compound_name,
            'pubchem_cid': phyto.pubchem_cid,
            'inchi_key': phyto.inchi_key,
            'smiles': phyto.canonical_smiles,
            'chemical_class': phyto.chemical_class,
            'tanimoto': round(r['similarity'], 4),
            'experiment_count': exps.count(),
            'synergy_count': exps.filter(interpretation='Synergy').count(),
        })

    return JsonResponse({
        'query_smiles': query_smiles,
        'fingerprint': 'Morgan/ECFP4 (radius 2, 2048-bit)',
        'metric': 'Tanimoto',
        'threshold': threshold,
        'count': len(data),
        'results': data,
    })


def api_docs(request):
    """API documentation page."""
    return render(request, 'synergy_data/api_docs.html')


# ==============================================================================
# HEALTH CHECK
# ==============================================================================

def health_check(request):
    """A simple view that proves the Django app is running."""
    return HttpResponse("Health Check OK. The Django application is running.", status=200)


# ==============================================================================
# ANALYTICS / DASHBOARD VIEW
# ==============================================================================

def analytics_page(request):
    """Public analytics dashboard with interactive visualizations."""
    experiments = SynergyExperiment.objects.select_related(
        'phytochemical', 'antibiotic', 'pathogen', 'source'
    )

    # ------------------------------------------------------------------
    # 3A. Synergy Interpretation Distribution (Donut Chart)
    # ------------------------------------------------------------------
    interpretation_counts = dict(
        experiments.values_list('interpretation')
        .annotate(count=Count('id'))
        .values_list('interpretation', 'count')
    )
    interpretation_data = {
        'labels': ['Synergy', 'Additive', 'Indifference', 'Antagonism'],
        'counts': [
            interpretation_counts.get('Synergy', 0),
            interpretation_counts.get('Additive', 0),
            interpretation_counts.get('Indifference', 0),
            interpretation_counts.get('Antagonism', 0),
        ],
        'colors': ['#0F0E47', '#272757', '#505081', '#8686AC'],
    }

    # ------------------------------------------------------------------
    # 3B. Experiments by ESKAPE Pathogen (Horizontal Bar Chart)
    # ------------------------------------------------------------------
    eskape_genera = ['Enterococcus', 'Staphylococcus', 'Klebsiella',
                     'Acinetobacter', 'Pseudomonas', 'Enterobacter']
    eskape_counts = []
    for genus in eskape_genera:
        cnt = experiments.filter(pathogen__genus__iexact=genus).count()
        eskape_counts.append(cnt)
    eskape_data = {
        'labels': eskape_genera,
        'counts': eskape_counts,
    }

    # ------------------------------------------------------------------
    # 3C. Top 10 Phytochemicals by Synergy Count
    # ------------------------------------------------------------------
    top_phyto = (
        experiments.filter(interpretation='Synergy')
        .values('phytochemical__compound_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    top_phyto_data = {
        'labels': [p['phytochemical__compound_name'] for p in top_phyto],
        'counts': [p['count'] for p in top_phyto],
    }

    # ------------------------------------------------------------------
    # 3D. Top 10 Antibiotics by Synergy Count
    # ------------------------------------------------------------------
    top_abx = (
        experiments.filter(interpretation='Synergy')
        .values('antibiotic__antibiotic_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    top_abx_data = {
        'labels': [a['antibiotic__antibiotic_name'] for a in top_abx],
        'counts': [a['count'] for a in top_abx],
    }

    # ------------------------------------------------------------------
    # 3E. Publication Year Trend (Line Chart)
    # ------------------------------------------------------------------
    year_trend = (
        experiments.filter(source__publication_year__isnull=False)
        .values('source__publication_year')
        .annotate(count=Count('id'))
        .order_by('source__publication_year')
    )
    year_data = {
        'labels': [y['source__publication_year'] for y in year_trend],
        'counts': [y['count'] for y in year_trend],
    }

    # ------------------------------------------------------------------
    # 3F. FIC Index Distribution (Histogram)
    # ------------------------------------------------------------------
    fic_values = list(
        experiments.filter(fic_index__isnull=False)
        .values_list('fic_index', flat=True)
    )
    fic_data = {
        'values': [float(v) for v in fic_values],
    }

    # ------------------------------------------------------------------
    # 3G. Synergy Heatmap (Phytochemical x Antibiotic)
    # ------------------------------------------------------------------
    top15_phyto = (
        experiments.values('phytochemical__id', 'phytochemical__compound_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:15]
    )
    top10_abx = (
        experiments.values('antibiotic__id', 'antibiotic__antibiotic_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    phyto_ids = [p['phytochemical__id'] for p in top15_phyto]
    abx_ids = [a['antibiotic__id'] for a in top10_abx]
    phyto_names = [p['phytochemical__compound_name'] for p in top15_phyto]
    abx_names = [a['antibiotic__antibiotic_name'] for a in top10_abx]

    heatmap_qs = (
        experiments.filter(
            phytochemical__id__in=phyto_ids,
            antibiotic__id__in=abx_ids,
            fic_index__isnull=False,
        )
        .values('phytochemical__id', 'antibiotic__id')
        .annotate(avg_fic=Avg('fic_index'))
    )
    heatmap_lookup = {}
    for row in heatmap_qs:
        key = (row['phytochemical__id'], row['antibiotic__id'])
        heatmap_lookup[key] = float(row['avg_fic'])

    heatmap_grid = []
    for p_id, p_name in zip(phyto_ids, phyto_names):
        row_data = {'name': p_name, 'cells': []}
        for a_id in abx_ids:
            avg_fic = heatmap_lookup.get((p_id, a_id))
            row_data['cells'].append(avg_fic)
        heatmap_grid.append(row_data)

    # ------------------------------------------------------------------
    # Summary stats
    # ------------------------------------------------------------------
    total_experiments = experiments.count()
    total_synergy = interpretation_counts.get('Synergy', 0)
    total_phytochemicals = Phytochemical.objects.count()
    total_antibiotics = Antibiotic.objects.count()
    total_pathogens = Pathogen.objects.count()
    total_sources = Source.objects.count()

    context = {
        'interpretation_json': json.dumps(interpretation_data),
        'eskape_json': json.dumps(eskape_data),
        'top_phyto_json': json.dumps(top_phyto_data),
        'top_abx_json': json.dumps(top_abx_data),
        'year_json': json.dumps(year_data),
        'fic_json': json.dumps(fic_data),
        'heatmap_grid': heatmap_grid,
        'heatmap_abx_names': abx_names,
        'total_experiments': total_experiments,
        'total_synergy': total_synergy,
        'total_phytochemicals': total_phytochemicals,
        'total_antibiotics': total_antibiotics,
        'total_pathogens': total_pathogens,
        'total_sources': total_sources,
    }
    return render(request, 'synergy_data/analytics.html', context)
